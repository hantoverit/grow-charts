#!/usr/bin/env python3
"""
freight_update.py — Automated weekly freight dashboard data update

Extracts data from a HANT-YYYY-MM-DD-W.XLS file and injects it into all 8
freight dashboard HTML widgets. Runs validation, reports results.

Usage:
    python freight_update.py HANT-2026-03-16-W.XLS

Requirements:
    - Python 3.8+
    - openpyxl (pip install openpyxl)
    - All 8 widget HTML files in the same directory (or specify --widgets-dir)
    - LibreOffice installed (for XLS→XLSX conversion)

The script:
    1. Converts XLS to XLSX via LibreOffice
    2. Extracts data from all relevant sheets
    3. Injects new week data into each widget's data structures
    4. Validates all widgets pass the audit checklist
    5. Reports results
"""

import argparse
import json
import os
import re
import subprocess
import sys
import tempfile
from collections import defaultdict
from pathlib import Path

# ═══════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════

WIDGET_FILES = [
    'freight-kpi-widget.html',
    'freight-carrier-bump.html',
    'freight-costdist-treemap.html',
    'freight-carrier-accessorial.html',
    'freight-accessorial-heatmap.html',
    'freight-savings-waterfall.html',
    'freight-surcharge-slope.html',
    'freight-shipment-bubble.html',
]

# Week label prefix: some widgets use "W" prefix, others don't
NO_PREFIX = {'freight-kpi-widget.html', 'freight-carrier-bump.html',
             'freight-costdist-treemap.html', 'freight-carrier-accessorial.html',
             'freight-accessorial-heatmap.html'}
W_PREFIX = {'freight-savings-waterfall.html', 'freight-surcharge-slope.html',
            'freight-shipment-bubble.html'}


def parse_week_label(filename):
    """Extract week label from HANT-YYYY-MM-DD-W.XLS filename."""
    m = re.match(r'HANT-\d{4}-(\d{2})-(\d{2})-W\.XLS', filename, re.IGNORECASE)
    if not m:
        sys.exit(f"ERROR: Filename '{filename}' doesn't match HANT-YYYY-MM-DD-W.XLS pattern")
    month, day = int(m.group(1)), int(m.group(2))
    return f"{month}-{day}"


def convert_xls(xls_path, tmp_dir):
    """Convert XLS to XLSX via LibreOffice."""
    print(f"  Converting {xls_path} to XLSX...")
    result = subprocess.run(
        ['libreoffice', '--headless', '--convert-to', 'xlsx', str(xls_path), '--outdir', tmp_dir],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        sys.exit(f"ERROR: LibreOffice conversion failed:\n{result.stderr}")
    xlsx_name = Path(xls_path).stem + '.xlsx'
    xlsx_path = os.path.join(tmp_dir, xlsx_name)
    if not os.path.exists(xlsx_path):
        sys.exit(f"ERROR: Expected converted file {xlsx_path} not found")
    return xlsx_path


# ═══════════════════════════════════════════════════════════
# DATA EXTRACTION
# ═══════════════════════════════════════════════════════════

def extract_all(xlsx_path, week_label):
    """Extract all data needed for 8 widgets from one weekly XLS file."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    data = {}

    # ── M1 KPI: Locn sheet ──
    ws = wb['Locn']
    data['kpi'] = {
        'totalShipments': ws.cell(2, 3).value or 0,      # TotalBills
        'avgPaidPerShipment': round(ws.cell(2, 8).value or 0, 2),  # AvgCost
        'totalFreightPaid': round(ws.cell(2, 5).value or 0, 2),    # TotalPaid
        'avgWeightPerShipment': ws.cell(2, 7).value or 0,  # AvgWgt
        'totalWeight': ws.cell(2, 4).value or 0,           # TotalWgt
    }

    # ── M2/3 Bump + name map: Carr sheet ──
    ws = wb['Carr']
    data['carriers'] = {}
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(r, 1).value
        if not cid:
            continue
        data['carriers'][cid] = {
            'name_raw': (ws.cell(r, 2).value or '').title().rstrip('.'),
            'bills': ws.cell(r, 4).value or 0,
            'paid': round(ws.cell(r, 6).value or 0, 2),
            'avg_cost': round(ws.cell(r, 8).value or 0, 2),
        }

    # ── M4 Treemap: CostDist sheet ──
    ws = wb['CostDist']
    data['costdist'] = {}
    for r in range(2, ws.max_row + 1):
        acct = ws.cell(r, 1).value
        paid = ws.cell(r, 2).value or 0
        if acct:
            data['costdist'][acct] = round(paid, 2)

    # ── M7 Waterfall: FrtSvgs sheet ──
    ws = wb['FrtSvgs']
    svgs_items = []
    for r in range(2, ws.max_row + 1):
        carr = ws.cell(r, 2).value
        if not carr:
            continue
        svgs_items.append({
            'carr': carr,
            'pro': str(ws.cell(r, 3).value or ''),
            'paid': round(ws.cell(r, 8).value or 0, 2),
            'billed': round(ws.cell(r, 9).value or 0, 2),
            'auditSvgs': round(ws.cell(r, 10).value or 0, 2),
            'dupSvgs': round(ws.cell(r, 11).value or 0, 2),
            'desc': str(ws.cell(r, 12).value or ''),
        })

    # Aggregate waterfall data
    carr_agg = defaultdict(lambda: {'billed': 0, 'paid': 0, 'auditSvgs': 0, 'dupSvgs': 0})
    for item in svgs_items:
        c = carr_agg[item['carr']]
        c['billed'] += item['billed']
        c['paid'] += item['paid']
        c['auditSvgs'] += item['auditSvgs']
        c['dupSvgs'] += item['dupSvgs']

    total_billed = sum(c['billed'] for c in carr_agg.values())
    total_paid = sum(c['paid'] for c in carr_agg.values())
    total_audit = sum(c['auditSvgs'] for c in carr_agg.values())
    total_dup = sum(c['dupSvgs'] for c in carr_agg.values())
    total_svgs = total_audit + total_dup
    savings_rate = round((total_svgs / total_billed * 100) if total_billed > 0 else 0, 1)

    data['waterfall'] = {
        'totals': {
            'billed': round(total_billed, 2), 'paid': round(total_paid, 2),
            'auditSvgs': round(total_audit, 2), 'dupSvgs': round(total_dup, 2),
            'totalSvgs': round(total_svgs, 2), 'savingsRate': savings_rate,
            'bills': len(svgs_items)
        },
        'carriers': [{'id': cid, 'billed': round(v['billed'], 2), 'paid': round(v['paid'], 2),
                       'auditSvgs': round(v['auditSvgs'], 2), 'dupSvgs': round(v['dupSvgs'], 2)}
                      for cid, v in sorted(carr_agg.items(), key=lambda x: -x[1]['billed'])],
        'items': svgs_items,
    }

    # ── M5/M6 Accessorials: Accessorials sheet ──
    ws = wb['Accessorials']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    acc_cols = {}
    for i, h in enumerate(headers):
        if i >= 9 and h:
            acc_cols[h] = i + 1

    # Type-first (M6) and Carrier-first (M5)
    hm_data = defaultdict(lambda: defaultdict(float))
    hm_detail = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {'s': 0, 'b': 0})))
    ca_data = defaultdict(float)
    ca_detail = defaultdict(lambda: defaultdict(lambda: {'s': 0, 'b': 0}))

    for r in range(2, ws.max_row + 1):
        carr = ws.cell(r, 2).value
        if not carr:
            continue
        for atype, col in acc_cols.items():
            val = ws.cell(r, col).value
            if val and val > 0:
                # M6 heatmap: type → week → total, type → week → carrier
                hm_data[atype][week_label] += val
                hm_detail[atype][week_label][carr]['s'] = round(hm_detail[atype][week_label][carr]['s'] + val, 2)
                hm_detail[atype][week_label][carr]['b'] += 1
                # M5 carrier acc: carrier → total, carrier → type
                ca_data[carr] += val
                ca_detail[carr][atype]['s'] = round(ca_detail[carr][atype]['s'] + val, 2)
                ca_detail[carr][atype]['b'] += 1

    # Round hm_data
    for atype in hm_data:
        for wk in hm_data[atype]:
            hm_data[atype][wk] = round(hm_data[atype][wk], 2)

    data['heatmap'] = {'data': dict(hm_data), 'detail': {t: {w: dict(c) for w, c in wks.items()} for t, wks in hm_detail.items()}}
    data['carrier_acc'] = {
        'data': {carr: round(total, 2) for carr, total in ca_data.items()},
        'detail': {carr: {atype: dict(d) for atype, d in types.items()} for carr, types in ca_detail.items()},
    }

    # ── M8 Slope: FecChgs + UpsAccs ──
    FEDEX_SURCHARGE_MAP = {
        'Delivery Area Surcharge': 'del_area', 'Delv Area Surchg - Com Express': 'del_area',
        'Delv Area Surchg - Res Express': 'del_area',
        'Res Delivery Gnd': 'res_del', 'Res Delivery Chg': 'res_del', 'Home Delivery Res Chg': 'res_del',
        'Late Fee': 'late_fee',
        'Oversize Extra Service Fee': 'oversize',
        'Additional Handling Chg': 'addl_hand', 'Addtl Handling - Wgt (Expr)': 'addl_hand',
        'ADDTL HANDLING CHG DIMS - GND': 'addl_dims',
        'ADDTL HANDLING CHG WGT - GND': 'addl_wgt',
        'Original Customs Duties': 'customs', 'Customs Entry Fee': 'customs',
        'Duty/Tax Adv. Fee': 'customs', 'Export Clearance': 'customs',
        'Goods & Service Tax Duty': 'customs', 'GST on Advance Svc Fees': 'customs',
        'Harmonized Sales Tax': 'customs', 'BRITISH COLUMBIA SALES TAX': 'customs',
        'Fuel Surcharge': 'fuel',
        'Third Party Billing Surcharge': '3p_bill',
        'Addr Corr Chg Gnd': 'addr_corr',
        'Peak Surcharge': 'peak', 'Surcharge DOI': 'peak',
        'SmartPost Non-Machineable Chg': 'peak',
        'Declared Value Gnd': 'declared',
        'Haz Material Gnd': 'haz_mat',
        'Direct Signature Required': 'declared',
        'Return Manager Chg': 'addr_corr',
    }
    UPS_SURCHARGE_MAP = {
        'DELIVERY AREA SURCHARGE': 'del_area',
        'RESIDENTIAL SURCHARGE': 'res_del',
        'ADDITIONAL HANDLING': 'addl_hand',
        'FUEL SURCHARGE': 'fuel',
        'DECLARED VALUE': 'declared',
        'THIRD PARTY COLLECT FEE': '3p_bill',
    }

    slope_data = defaultdict(lambda: {'fedex': 0, 'ups': 0})

    ws = wb['FecChgs']
    for r in range(2, ws.max_row + 1):
        desc = ws.cell(r, 6).value or ''
        amt = ws.cell(r, 5).value or 0
        if desc and amt != 0:
            cat_id = FEDEX_SURCHARGE_MAP.get(desc)
            if cat_id:
                slope_data[cat_id]['fedex'] += amt

    ws = wb['UpsAccs']
    for r in range(2, ws.max_row + 1):
        desc = ws.cell(r, 3).value or ''
        amt = ws.cell(r, 5).value or 0
        if desc and amt != 0:
            cat_id = UPS_SURCHARGE_MAP.get(desc)
            if cat_id:
                slope_data[cat_id]['ups'] += amt

    data['slope'] = {cat: {'fedex': round(v['fedex'], 2), 'ups': round(v['ups'], 2)}
                     for cat, v in slope_data.items()}

    # ── M9 Bubble: FecPkgs + UpsPkgs + Bills ──
    FEDEX_SVC_MAP = {
        'CG': 'fedex_ground', 'R5': 'fedex_home_delivery', 'DS': 'fedex_2day',
        'AE': 'fedex_intl_economy', 'ON': 'fedex_first_overnight',
        'CX': 'fedex_express_saver', 'ST': 'fedex_std_overnight',
        'NM': 'fedex_intl_next_flight', 'ES': 'fedex_express_saver',
    }
    UPS_SVC_MAP = {
        'GND': 'ups_ground', 'NDA': 'ups_next_day_air',
        'NDA AM': 'ups_nda_early', '2DA': 'ups_2day_air_am',
    }

    bubble_data = defaultdict(lambda: {'count': 0, 'spend': 0})

    ws = wb['FecPkgs']
    for r in range(2, ws.max_row + 1):
        svc_code = ws.cell(r, 16).value or ''
        paid = ws.cell(r, 18).value or 0
        svc_id = FEDEX_SVC_MAP.get(svc_code)
        if svc_id:
            bubble_data[svc_id]['count'] += 1
            bubble_data[svc_id]['spend'] += paid

    ws = wb['UpsPkgs']
    for r in range(2, ws.max_row + 1):
        svc_code = ws.cell(r, 7).value or ''
        paid = ws.cell(r, 10).value or 0
        svc_id = UPS_SVC_MAP.get(svc_code)
        if svc_id:
            bubble_data[svc_id]['count'] += 1
            bubble_data[svc_id]['spend'] += paid

    ws = wb['Bills']
    for r in range(2, ws.max_row + 1):
        mode = ws.cell(r, 3).value or ''
        carr = ws.cell(r, 2).value or ''
        paid = ws.cell(r, 10).value or 0
        if carr.startswith(('FX', 'FE', 'FEC', 'UP', 'UPS')):
            continue
        if mode == 'LTL':
            bubble_data['ltl_freight']['count'] += 1
            bubble_data['ltl_freight']['spend'] += paid
        elif mode == 'TL':
            bubble_data['truckload']['count'] += 1
            bubble_data['truckload']['spend'] += paid

    data['bubble'] = {sid: {'count': v['count'], 'spend': round(v['spend'], 2)}
                      for sid, v in bubble_data.items()}

    return data


# ═══════════════════════════════════════════════════════════
# CARRIER NAME MAP
# ═══════════════════════════════════════════════════════════

def build_name_map(html):
    """Extract existing carrier name conventions from bump chart DATA."""
    names = {}
    for m in re.finditer(r'"(\w+\*?)":\s*\{\s*"name":\s*"([^"]+)"', html):
        cid, name = m.group(1), m.group(2)
        if cid not in names:
            names[cid] = name
    return names


# ═══════════════════════════════════════════════════════════
# INJECTION — per widget
# ═══════════════════════════════════════════════════════════

def inject_weeks_array(html, week_label, has_prefix):
    """Add new week to WEEKS array if not already present."""
    prefix = 'W' if has_prefix else ''
    wk_str = f'{prefix}{week_label}'

    # Match both single and double quote styles
    m = re.search(r'((?:const|var)\s+WEEKS\s*=\s*\[)(.*?)(\])', html)
    if not m:
        return html, False

    existing = m.group(2)
    if wk_str in existing:
        return html, False  # Already present

    # Add new week
    quote = "'" if "'" in existing else '"'
    new_entry = f',{quote}{wk_str}{quote}'
    new_weeks = m.group(1) + existing + new_entry + m.group(3)
    html = html[:m.start()] + new_weeks + html[m.end():]
    return html, True


def inject_kpi(html, week_label, kpi_data):
    """Inject KPI data into M1."""
    # Find the data object and add new week
    pattern = r'(const data=\{)(.*?)(\};)'
    m = re.search(pattern, html, re.DOTALL)
    if not m:
        return html

    existing = m.group(2)
    if f'"{week_label}"' in existing:
        return html

    new_entry = (f',\n  "{week_label}":{{totalShipments:{kpi_data["totalShipments"]},'
                 f'avgPaidPerShipment:{kpi_data["avgPaidPerShipment"]},'
                 f'totalFreightPaid:{kpi_data["totalFreightPaid"]},'
                 f'avgWeightPerShipment:{kpi_data["avgWeightPerShipment"]},'
                 f'totalWeight:{kpi_data["totalWeight"]}}}')
    html = html[:m.end(2)] + new_entry + html[m.end(2):]
    return html


def inject_bump(html, week_label, carriers, name_map):
    """Inject bump chart carrier data."""
    # DATA is {week: {carrier: {name, bills, paid, avg_cost}}}
    m = re.search(r'(const DATA=\{)(.*?)(\};)', html, re.DOTALL)
    if not m or f'"{week_label}"' in m.group(2):
        return html

    entries = []
    for cid, d in carriers.items():
        name = name_map.get(cid, d['name_raw'])
        entries.append(f'"{cid}":{{"name":"{name}","bills":{d["bills"]},"paid":{d["paid"]},"avg_cost":{d["avg_cost"]}}}')

    new_week = f',"{week_label}":{{{",".join(entries)}}}'
    html = html[:m.end(2)] + new_week + html[m.end(2):]
    return html


def inject_treemap(html, week_label, costdist):
    """Inject treemap CostDist data (unquoted JS keys)."""
    m = re.search(r'(const RAW=\{)(.*?)(\};)', html, re.DOTALL)
    if not m or f'"{week_label}"' in m.group(2):
        return html

    entries = ','.join(f'{acct}:{paid}' for acct, paid in costdist.items())
    new_week = f',\n  "{week_label}":{{{entries}}}'
    html = html[:m.end(2)] + new_week + html[m.end(2):]
    return html


def inject_heatmap(html, week_label, hm_data, hm_detail):
    """Inject heatmap accessorial data (type-first) using JSON parse/modify/serialize."""
    # ── DATA[type][week] = total ──
    data_m = re.search(r'const DATA=(\{.*?\});\s*const DETAIL', html, re.DOTALL)
    if not data_m:
        return html

    try:
        data_obj = json.loads(data_m.group(1))
    except json.JSONDecodeError:
        return html  # Can't parse — skip

    # Check if week already present in first type
    first_type = next(iter(data_obj), None)
    if first_type and week_label in data_obj[first_type]:
        return html  # Already injected

    # Add new week to every existing type
    for atype in data_obj:
        if atype in hm_data and week_label in hm_data[atype]:
            data_obj[atype][week_label] = hm_data[atype][week_label]
        else:
            data_obj[atype][week_label] = 0

    # Add new types not yet in data
    for atype, weeks in hm_data.items():
        if atype not in data_obj and week_label in weeks:
            data_obj[atype] = {week_label: weeks[week_label]}

    new_data = json.dumps(data_obj, separators=(',', ':'))
    # Re-format to match original style (spaces after colons for readability)
    new_data = new_data.replace(',"', ', "').replace('{', '{ ').replace('}', ' }')
    # Actually keep it compact like original
    new_data = json.dumps(data_obj)
    html = html[:data_m.start(1)] + new_data + html[data_m.end(1):]

    # ── DETAIL[type][week][carrier] = {s, b} ──
    detail_m = re.search(r'const DETAIL=(\{.*?\});\s', html, re.DOTALL)
    if detail_m:
        try:
            detail_obj = json.loads(detail_m.group(1))
        except json.JSONDecodeError:
            return html

        # Add new week's carrier data for each type
        for atype, weeks in hm_detail.items():
            if week_label not in weeks:
                continue
            if atype not in detail_obj:
                detail_obj[atype] = {}
            detail_obj[atype][week_label] = weeks[week_label]

        # For types that exist but have no data this week, add empty {}
        for atype in detail_obj:
            if week_label not in detail_obj[atype]:
                detail_obj[atype][week_label] = {}

        new_detail = json.dumps(detail_obj)
        html = html[:detail_m.start(1)] + new_detail + html[detail_m.end(1):]

    return html


def inject_carrier_acc(html, week_label, ca_data, ca_detail, carriers, name_map):
    """Inject carrier accessorial data (carrier-first) using JSON parse/modify/serialize."""
    # ── DATA[carrier] = {name, week: total} ──
    data_m = re.search(r'const DATA=(\{.*?\});\s*const DETAIL', html, re.DOTALL)
    if not data_m:
        return html

    try:
        data_obj = json.loads(data_m.group(1))
    except json.JSONDecodeError:
        return html

    # Check if week already present
    first_carr = next(iter(data_obj), None)
    if first_carr and week_label in data_obj[first_carr]:
        return html

    # Add new week to every existing carrier
    for carr in data_obj:
        data_obj[carr][week_label] = round(ca_data.get(carr, 0), 2)

    # Add new carriers not yet in data
    for carr, total in ca_data.items():
        if carr not in data_obj:
            name = name_map.get(carr, carriers.get(carr, {}).get('name_raw', carr))
            data_obj[carr] = {"name": name, week_label: round(total, 2)}

    new_data = json.dumps(data_obj)
    html = html[:data_m.start(1)] + new_data + html[data_m.end(1):]

    # ── DETAIL[carrier][type][week] = {s, b} ──
    detail_m = re.search(r'const DETAIL=(\{.*?\});\s', html, re.DOTALL)
    if detail_m:
        try:
            detail_obj = json.loads(detail_m.group(1))
        except json.JSONDecodeError:
            return html

        for carr, types in ca_detail.items():
            if carr not in detail_obj:
                detail_obj[carr] = {}
            for atype, d in types.items():
                if atype not in detail_obj[carr]:
                    detail_obj[carr][atype] = {}
                detail_obj[carr][atype][week_label] = d

        new_detail = json.dumps(detail_obj)
        html = html[:detail_m.start(1)] + new_detail + html[detail_m.end(1):]

    return html


def inject_waterfall(html, week_label, wf_data):
    """Inject waterfall savings data."""
    wk_key = f'W{week_label}'
    m = re.search(r"(const DATA=\{)(.*?)(\};)", html, re.DOTALL)
    if not m or f'"{wk_key}"' in m.group(2):
        return html

    new_week = json.dumps({wk_key: wf_data}, indent=2)[1:-1]  # Strip outer braces
    html = html[:m.end(2)] + ',' + new_week + html[m.end(2):]
    return html


def inject_slope(html, week_label, slope_data):
    """Inject slope surcharge data into CATS array entries."""
    wk_key = f'W{week_label}'

    # For each category in CATS, find the object by id and add the week to fedex/ups
    for cat_id, vals in slope_data.items():
        # Find the category object: {id:"cat_id",...,fedex:{...},ups:{...}}
        # Add new week to fedex object
        fedex_pattern = rf'(id:"{re.escape(cat_id)}"[^}}]*?fedex:\{{[^}}]*)'
        fedex_m = re.search(fedex_pattern, html)
        if fedex_m and f'"{wk_key}"' not in fedex_m.group(1):
            html = html[:fedex_m.end(1)] + f',"{wk_key}":{vals["fedex"]}' + html[fedex_m.end(1):]

        # Add new week to ups object (re-search since html changed)
        ups_pattern = rf'(id:"{re.escape(cat_id)}"[^}}]*?ups:\{{[^}}]*)'
        ups_m = re.search(ups_pattern, html)
        if ups_m and f'"{wk_key}"' not in ups_m.group(1):
            html = html[:ups_m.end(1)] + f',"{wk_key}":{vals["ups"]}' + html[ups_m.end(1):]

    # For categories in CATS that have no data this week, add 0 values
    existing_cats = re.findall(r'id:"([^"]+)"', html)
    for cat_id in existing_cats:
        if cat_id not in slope_data:
            fedex_pattern = rf'(id:"{re.escape(cat_id)}"[^}}]*?fedex:\{{[^}}]*)'
            fedex_m = re.search(fedex_pattern, html)
            if fedex_m and f'"{wk_key}"' not in fedex_m.group(1):
                html = html[:fedex_m.end(1)] + f',"{wk_key}":0' + html[fedex_m.end(1):]
            ups_pattern = rf'(id:"{re.escape(cat_id)}"[^}}]*?ups:\{{[^}}]*)'
            ups_m = re.search(ups_pattern, html)
            if ups_m and f'"{wk_key}"' not in ups_m.group(1):
                html = html[:ups_m.end(1)] + f',"{wk_key}":0' + html[ups_m.end(1):]

    return html


def inject_bubble(html, week_label, bubble_data):
    """Inject bubble service level data into SVCS array entries."""
    wk_key = f'W{week_label}'

    # For each service in SVCS, find by id and add the week data
    existing_svcs = re.findall(r'id:"([^"]+)"', html)

    for svc_id in existing_svcs:
        # Skip non-service IDs (e.g., from info card elements)
        if svc_id not in bubble_data and svc_id not in (
            'fedex_2day', 'fedex_express_saver', 'fedex_first_overnight', 'fedex_ground',
            'fedex_home_delivery', 'fedex_intl_economy', 'fedex_intl_next_flight',
            'fedex_std_overnight', 'ltl_freight', 'truckload',
            'ups_2day_air_am', 'ups_ground', 'ups_nda_early', 'ups_next_day_air'
        ):
            continue

        vals = bubble_data.get(svc_id, {'count': 0, 'spend': 0})
        new_data = f'"{wk_key}":{{"count":{vals["count"]},"spend":{vals["spend"]}}}'

        # Find the service object and add before the closing }
        # Pattern: id:"svc_id",label:...,carrier:...,"W3-2":{...}}
        # We need to find the last } of this service entry
        svc_start = html.find(f'id:"{svc_id}"')
        if svc_start < 0:
            continue

        # Check if week already present
        # Find the end of this service object (next },{id: or }]; )
        search_from = svc_start
        next_svc = html.find('},\n{id:', search_from + 1)
        next_end = html.find('}\n];', search_from + 1)
        if next_svc < 0:
            next_svc = float('inf')
        if next_end < 0:
            next_end = float('inf')
        svc_end = min(next_svc, next_end)

        svc_block = html[svc_start:int(svc_end) + 1]
        if f'"{wk_key}"' in svc_block:
            continue

        # Insert before the closing } of this service object
        insert_pos = int(svc_end)
        html = html[:insert_pos] + ',' + new_data + html[insert_pos:]

    return html


# ═══════════════════════════════════════════════════════════
# VALIDATION
# ═══════════════════════════════════════════════════════════

def validate_widget(html, filename, week_label):
    """Run audit checks on a single widget."""
    issues = []
    has_prefix = filename in W_PREFIX
    prefix = 'W' if has_prefix else ''
    wk_str = f'{prefix}{week_label}'

    # Check WEEKS array contains new week
    if wk_str not in html:
        issues.append(f"Week '{wk_str}' not found in file")

    # Check no hardcoded week buttons in HTML section
    script_start = html.find('<script')
    html_part = html[:script_start] if script_start > -1 else html
    hardcoded_btns = re.findall(r'onclick="setWeek\(\'[\d-]+\'\)"', html_part)
    if hardcoded_btns:
        issues.append(f"Hardcoded week buttons in HTML: {len(hardcoded_btns)}")

    # Check color arrays
    for pat in ['WCOL_L', 'WCOL_D', 'CCOL_L', 'CCOL_D']:
        m = re.search(rf'{pat}=\[(.*?)\]', html)
        if m and m.group(1).count('#') < 4:
            issues.append(f"{pat} has fewer than 4 colors")

    # Check no hardcoded "N-Week Total"
    if re.search(r'\d-Week Total', html) and 'WEEKS.length' not in html:
        issues.append("Hardcoded N-Week Total string")

    return issues


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description='Update freight dashboard with new weekly data')
    parser.add_argument('xls_file', help='Path to HANT-YYYY-MM-DD-W.XLS file')
    parser.add_argument('--widgets-dir', default='.', help='Directory containing widget HTML files (default: current dir)')
    parser.add_argument('--dry-run', action='store_true', help='Extract and validate without modifying files')
    args = parser.parse_args()

    xls_path = Path(args.xls_file)
    widgets_dir = Path(args.widgets_dir)

    if not xls_path.exists():
        sys.exit(f"ERROR: File not found: {xls_path}")

    # Verify all widget files exist
    missing = [f for f in WIDGET_FILES if not (widgets_dir / f).exists()]
    if missing:
        sys.exit(f"ERROR: Missing widget files in {widgets_dir}: {', '.join(missing)}")

    week_label = parse_week_label(xls_path.name)
    print(f"\n{'='*60}")
    print(f"FREIGHT DASHBOARD UPDATE — W{week_label}")
    print(f"{'='*60}")
    print(f"  Source: {xls_path.name}")
    print(f"  Week label: {week_label} (buttons: W{week_label})")
    print(f"  Widgets dir: {widgets_dir}")

    # Step 1: Convert XLS
    print(f"\n[1/5] Converting XLS to XLSX...")
    with tempfile.TemporaryDirectory() as tmp_dir:
        xlsx_path = convert_xls(str(xls_path), tmp_dir)

        # Step 2: Extract data
        print(f"[2/5] Extracting data from all sheets...")
        data = extract_all(xlsx_path, week_label)
        print(f"  KPI: {data['kpi']['totalShipments']} shipments, ${data['kpi']['totalFreightPaid']:,.2f} freight")
        print(f"  Carriers: {len(data['carriers'])} in Carr sheet")
        print(f"  CostDist: {len(data['costdist'])} accounts")
        print(f"  Waterfall: {data['waterfall']['totals']['bills']} savings items, ${data['waterfall']['totals']['totalSvgs']:,.2f} total savings")
        print(f"  Accessorials: {len(data['heatmap']['data'])} types, {len(data['carrier_acc']['data'])} carriers")

    # Step 3: Build carrier name map from existing bump chart
    print(f"\n[3/5] Building carrier name map...")
    bump_html = (widgets_dir / 'freight-carrier-bump.html').read_text()
    name_map = build_name_map(bump_html)
    new_carriers = set(data['carriers'].keys()) - set(name_map.keys())
    if new_carriers:
        print(f"  New carriers (not in prior weeks): {new_carriers}")
        for c in new_carriers:
            name_map[c] = data['carriers'][c]['name_raw']
    print(f"  Name map: {len(name_map)} carriers")

    if args.dry_run:
        print(f"\n[DRY RUN] Extraction complete. No files modified.")
        return

    # Step 4: Inject into each widget
    print(f"\n[4/5] Injecting W{week_label} into all widgets...")
    results = {}
    for fname in WIDGET_FILES:
        fpath = widgets_dir / fname
        html = fpath.read_text()
        has_prefix = fname in W_PREFIX
        original_len = len(html)

        # Inject WEEKS array
        html, added = inject_weeks_array(html, week_label, has_prefix)

        # Widget-specific injection
        if fname == 'freight-kpi-widget.html':
            html = inject_kpi(html, week_label, data['kpi'])
        elif fname == 'freight-carrier-bump.html':
            html = inject_bump(html, week_label, data['carriers'], name_map)
        elif fname == 'freight-costdist-treemap.html':
            html = inject_treemap(html, week_label, data['costdist'])
        elif fname == 'freight-accessorial-heatmap.html':
            html = inject_heatmap(html, week_label, data['heatmap']['data'], data['heatmap']['detail'])
        elif fname == 'freight-carrier-accessorial.html':
            html = inject_carrier_acc(html, week_label, data['carrier_acc']['data'],
                                       data['carrier_acc']['detail'], data['carriers'], name_map)
        elif fname == 'freight-savings-waterfall.html':
            html = inject_waterfall(html, week_label, data['waterfall'])

        elif fname == 'freight-surcharge-slope.html':
            html = inject_slope(html, week_label, data['slope'])
        elif fname == 'freight-shipment-bubble.html':
            html = inject_bubble(html, week_label, data['bubble'])

        delta = len(html) - original_len
        fpath.write_text(html)
        results[fname] = 'OK' if delta > 0 else 'SKIP (already present)'
        print(f"    ✅ {fname}: {'+' if delta > 0 else ''}{delta} bytes")

    # Step 5: Validate
    print(f"\n[5/5] Validating all widgets...")
    all_clean = True
    for fname in WIDGET_FILES:
        html = (widgets_dir / fname).read_text()
        issues = validate_widget(html, fname, week_label)
        if issues:
            all_clean = False
            print(f"    ❌ {fname}:")
            for i in issues:
                print(f"       {i}")
        else:
            print(f"    ✅ {fname}")

    # Summary
    print(f"\n{'='*60}")
    print(f"RESULTS")
    print(f"{'='*60}")
    for fname, status in results.items():
        print(f"  {fname}: {status}")
    print(f"\n  Validation: {'ALL CLEAN ✅' if all_clean else 'ISSUES FOUND ❌'}")

    print(f"\n  Next steps:")
    print(f"    1. git add -A && git commit -m 'W{week_label} data update'")
    print(f"    2. git push origin main")
    from datetime import date
    today = date.today().strftime('%Y%m%d')
    print(f"    3. Update Grow embed URLs: append ?v={today}")
    print()


if __name__ == '__main__':
    main()
